import datetime
import random
import requests
from urllib.parse import urljoin
from urllib.request import urlopen
from bs4 import BeautifulSoup

web_url = r'http://www.imooc.com'
html = urlopen(web_url)
bs = BeautifulSoup(html, 'html.parser')
random.seed(datetime.datetime.now().timestamp())


def getLinks(other_url):
    html = urlopen(web_url + f'{other_url}')
    bs = BeautifulSoup(html, 'html.parser')
    return bs.find('div', {'class': 'course-list'}).find_all(lambda tag: len(tag.attrs) == 6)


"""
遍历所有网站节点并保存(每种课程的网页)
"""
externalUrl = []
try:
    for i in range(1, 23):
        links = getLinks('/course/list?ct=1&page=' + str(i))

        for i in range(40):
            newArticle = links[random.randint(0, len(links) - 1)].attrs['href']
            externalUrl.append('http:'+newArticle)
except AttributeError:
    pass



print('externalurl scrapy is Done!')
class Course:
    def __init__(self, name, time, teacher, introduction, price, address):
        self.name = name
        self.time = time
        self.teacher = teacher
        self.introduction = introduction
        self.price = price
        self.address = address


class solution(object):

    def name(self, externalUrl):
        html = urlopen(externalUrl)
        bs = BeautifulSoup(html, 'html.parser')
        name = bs.find('div', {'class': 'path'}).find('span').getText()
        return name

    def time(self, externalUrl):
        html = urlopen(externalUrl)
        bs = BeautifulSoup(html, 'html.parser')
        try:
            time = bs.find('div', {'class': 'info-bar clearfix'}).findAll('span', {'class': 'nodistance'})[1].text
        except AttributeError:
            #print('Error')
            #print(externalUrl)

            time=bs.findAll('span',{'class':'meta-value'})[1].text
            #print(f'time is {time}')
        return time

    def teacher(self, externalUrl):
        html = urlopen(externalUrl)
        bs = BeautifulSoup(html, 'html.parser')
        try:
            teacher = bs.find('div', {'class': 't-name'}).get_text()
        except:
            #print('Error!')
            #print(externalUrl)
            teacher=bs.find('span',{'class':'tit'}).find('a').text
            #print(f'teacher is {teacher}')
        return teacher

    def introduction(self, externalUrl):
        html = urlopen(externalUrl)
        bs = BeautifulSoup(html, 'html.parser')
        try:
            introduction = bs.find('div', {'class': 'dec-box'}).get_text()
        except:
            #print('Error')
            introduction = bs.find('div', {"class":"course-description course-wrap"}).text
            #print(f'introdution is {introduction}')
        return introduction

    def price(self, externalUrl):
        html = urlopen(externalUrl)
        bs = BeautifulSoup(html, 'html.parser')
        try:
            price = bs.find('div', {'class': 'ori-price l'}).get_text().split()[0]
        except:
            price = '0'
        return price

    def address(self, externalUrl):
        return externalUrl


def writeDown():
    import openpyxl
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    try:
        worksheet.cell(row=1, column=1, value="name")
        worksheet.cell(row=1, column=2, value="time")
        worksheet.cell(row=1, column=3, value="teacher")
        worksheet.cell(row=1, column=4, value="introduction")
        worksheet.cell(row=1, column=5, value="price")
        worksheet.cell(row=1, column=6, value="address")
        Solution=solution()

        for i in range(2, len(externalUrl) + 1):
            Solution.cover(externalUrl[i - 1])
            worksheet.cell(row=i, column=1, value=Solution.name(externalUrl[i-1]))
            worksheet.cell(row=i, column=2, value=Solution.time(externalUrl[i-1]))
            worksheet.cell(row=i, column=3, value=Solution.teacher(externalUrl[i-1]))
            worksheet.cell(row=i, column=4, value=Solution.introduction(externalUrl[i-1]))
            worksheet.cell(row=i, column=5, value=Solution.price(externalUrl[i-1]))
            worksheet.cell(row=i, column=6, value=Solution.address(externalUrl[i-1]))

    except:
        pass
    finally:
        workbook.save("imooc.xlsx")


writeDown()
Web=r'http://www.imooc.com'
def cover(web_url,n):
    import datetime
    import random
    import re
    import requests
    from urllib.parse import urljoin
    from urllib.request import urlopen, Request
    from bs4 import BeautifulSoup
    import time
    header = {'User-Agent':
                  'Mozilla'}

    req = Request(web_url, headers=header)
    html = urlopen(req)
    bs = BeautifulSoup(html, 'html.parser')
    random.seed(datetime.datetime.now().timestamp())
    web_url = web_url + '/course/list?ct=1&page=' + str(n)
    bs = BeautifulSoup(urlopen(web_url), 'html.parser')
    cover = bs.findAll('div', {'class': 'img'})
    for c in cover:
        # print(type(c['style']))
        U = re.search(r"url\('?(.*?)'?\)", c['style']).group(1)
        # print(U.split('/')[-1])
        response = requests.get('https:' + U, header,verify=False)
        time.sleep(random.random())
        filename = "D:\\Developer environment\\Study\\scrapy\\scrapyImage\\" + U.split('/')[-1]
        # print(filename)
        with open(filename, 'wb') as f:
            f.write(response.content)
for i in range(1,24):
    cover(Web,i)